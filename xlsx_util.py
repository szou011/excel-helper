#!/usr/bin/env python3
# coding: utf-8
"""A excel utility library for easy excel mulnipulation.

This module is based on openpyxl and win32com package. Each package serves different purpose
(i.e.openpyxl works on mutiple platform incl. Linux/MacOS however win32com interacts with 
Excel through windows COM interface, there are also few features only avaialble in win32com)


The module includes two major classes:

OpenPyWorksheet: a wrapper of the OpenPyXlsx worksheet class
Win32Worksheet: a wrapper of the win32com Excel.Application.Workbook.Worksheet class

"""

# third party modules
from openpyxl import load_workbook
import win32com.client

# import built-in modules
import csv
import gc


class OpenPyWorksheet:
    def __init__(
        self,
        path,
        active_worksheet=None,
        header_row=None,
        header_columns=None,
        end_row=None,
    ):
        self.workbook = load_workbook(path)

        if not active_worksheet:
            self.worksheet = self.workbook.active
        else:
            assert active_worksheet in self.workbook.sheetnames
            self.worksheet = self.workbook[active_worksheet]

        self.header_row = header_row
        self.header_columns = header_columns
        self.end_row = end_row

    def _unmerge_cells(self):
        merged_cell_ranges_string = [
            str(merged) for merged in self.worksheet.merged_cells
        ]

        for range_string in merged_cell_ranges_string:
            self.worksheet.unmerge_cells(range_string=range_string)

    def _autoset_data_range(self):
        self._unmerge_cells()

        if not self.header_row:
            self.header_row = self._header_row()

        if not self.header_columns:
            self.header_columns = self._header_columns()

        if not self.end_row:
            self.end_row = self.worksheet.max_row

    def _header_row(self):
        for row in range(1, 11):
            for col in range(1, 11):
                if self.worksheet.cell(row=row, column=col).value:
                    return row
        return None

    def _header_columns(self):
        return [
            header_column
            for header_column in range(1, self.worksheet.max_column + 1)
            if self.worksheet.cell(row=self.header_row, column=header_column).value
        ]

    @property
    def header(self):
        self._autoset_data_range()

        return [
            self.worksheet.cell(row=self.header_row, column=col).value
            for col in self.header_columns
        ]

    @property
    def data(self):
        self._autoset_data_range()

        for row in range(self.header_row, self.end_row + 1):
            row_data = [
                self.worksheet.cell(row=row, column=col).value
                for col in self.header_columns
            ]

            yield row_data

    def export_to_csv(self, path):
        with open(path, "w", newline="") as csvfile:
            csv_writer = csv.writer(csvfile)
            csv_writer.writerows(self.data)


class Win32Worksheet:
    def __init__(self, path, active_worksheet=None):

        # using gencache.EnsureDispatch to create a static proxy
        # make sure to run makepy using scripts below and close Excel applicaiton (if it's open):
        # from win32com.client import makepy
        # makepy.main()

        self.excel = win32com.client.gencache.EnsureDispatch('Excel.Application')

        # disable Excel UI, auto update and events
        # most of the time the excel app is only required at the back end

        self.excel.Visible = False 
        self.excel.ScreenUpdating = False
        self.excel.DisplayStatusBar = False
        self.EnableEvents = False


        self.workbook = self.excel.Workbooks.Open(path)

        if active_worksheet:
            self.worksheet = self.workbook.Worksheets(active_worksheet)
        else:
            self.worksheet = self.workbook.ActiveSheet

    def __enter__(self):
        return self


    def print_to_pdf(self, output_path):
        """A wrapper of ExportAsFixedFormat method.

        Parameters
        ----------
        output_path
            A string that indicates the name of the file to be saved.

        Returns
        -------
        None

        """

        # ExportAsFixedFormat takes XlFixedFormatType enumeration
        # 0 is for PDF, 1 is for XPS

        if not isinstance(output_path, str):
            output_path = str(output_path)

        self.worksheet.ExportAsFixedFormat(0, output_path)

        return None


    def __exit__(self, type, value, traceback):
        # make sure the excel Applicaion quits every time by using the context manager

        self.workbook.Close()
        # may accidentally close all other excel self.excel.Quit()
        self.excel = None
        self.workbook = None
        self.worksheet = None
        gc.collect()
